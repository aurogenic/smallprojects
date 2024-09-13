import requests
# import json
from openpyxl import Workbook, load_workbook


def main():
    
    create_excel_file("credits.xlsx", ['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'] )
    create_excel_file("scores.xlsx", ['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'] )
    cookies = {
        'JSESSIONID': '05B59C1337D7D205F245D6A6F8E6BD0E'
    }
    headers = {
        "Host": "klsgroup.dhi-edu.com",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://klsgroup.dhi-edu.com/",
        "Content-Type": "application/json",
        "Authorization": "Bearer eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJRTEtLRmhId3F1bHFkUHNSQl9WaG1pT3RHRlRaR01mYW9sT1l4ZzFWVm9FIn0.eyJqdGkiOiJkODg4MDE2OC1mZTJiLTRhOWUtYmY5NC0xZTZlNDkwOTZlZTAiLCJleHAiOjE3MTY5MzcyNzAsIm5iZiI6MCwiaWF0IjoxNzE2OTE5MjcyLCJpc3MiOiJodHRwczovL2F1dGguZGhpLWVkdS5jb20vYXV0aC9yZWFsbXMva2xzZ3JvdXAiLCJhdWQiOiJhY2NvdW50Iiwic3ViIjoiZjoyZGFlMzNkOC1mMTIwLTQ3MzAtYjM4My1jZGQyNDM1NjNjZGI6Z2l0MjNjczA1Ny10IiwidHlwIjoiQmVhcmVyIiwiYXpwIjoia2xzZ3JvdXBfZ2l0Iiwibm9uY2UiOiI5NDE3NjBlZi0xYWRhLTQ3NDItODM0ZS05OTNhN2NjNWViMDIiLCJhdXRoX3RpbWUiOjE3MTY5MTkyNzAsInNlc3Npb25fc3RhdGUiOiJiZWUzNTEyOS05MjQyLTRlMTktODNiMy0xNzNjYjZkMjZkNzciLCJhY3IiOiIxIiwiYWxsb3dlZC1vcmlnaW5zIjpbIioiXSwicmVhbG1fYWNjZXNzIjp7InJvbGVzIjpbIm9mZmxpbmVfYWNjZXNzIiwidW1hX2F1dGhvcml6YXRpb24iXX0sInJlc291cmNlX2FjY2VzcyI6eyJrbHNncm91cF9naXQiOnsicm9sZXMiOlsiU1RVREVOVCJdfSwiYWNjb3VudCI6eyJyb2xlcyI6WyJtYW5hZ2UtYWNjb3VudCIsIm1hbmFnZS1hY2NvdW50LWxpbmtzIiwidmlldy1wcm9maWxlIl19fSwic2NvcGUiOiJvcGVuaWQgZW1haWwgcHJvZmlsZSIsImVtYWlsX3ZlcmlmaWVkIjpmYWxzZSwicHJlZmVycmVkX3VzZXJuYW1lIjoiZ2l0MjNjczA1Ny10IiwiZW1haWwiOiJnaXQyM2NzMDU3LXQifQ.FDAIbNUqSDkqcAbz80WXZHw5P5191wo6wgEo9etEJ9tzxs3nJWn9JrvnfmdDr33yliUSJdBy56dYtqPWvKYac4pTACW--bpG89N85SqMYvzxbkzpx7v4cKVa4JrND4doxTqNFrDnsgLK4UEngJWdAEAhr5-tFf6er28j7gNmpvQeJ0tTlNDnK20BPtm77IKZ3Rd9I8Sa-MzWyhmPRDlpsNtERbICRFprNVV-3yYHh1R2VMH0BOvUWovP8qIhw6vZe11eX0h9hWngmW4z9rz9gMwwwAaGNcBKLOSHixo0GCAJSgitA62cbDDrvu5JKNCaKtLvYhIH-tTj0ZpWOHSP4g",
        "tenant-id": "klsgroup_git",
        "academic-type": "COLLEGE",
        "x-real-ip": "localhost",
        "Connection": "keep-alive",
        "Cookie": "JSESSIONID=05B59C1337D7D205F245D6A6F8E6BD0E",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin"
    }
    for i in range(9):
        try:
            USN = 'GIT23CS00' + str(i) + '-T'
            url = "https://klsgroup.dhi-edu.com/dhiapiserver/api/university-exam/score/students/64e4697835a997bc3f6d305a?academicyear=2023-24&degree=BE&departmentid=CS&termNumber=1&scheme=Scheme%202022&examid=DECEMBER_FEBRUARY_2023-24&usn=" + USN
            params = {
                "academicyear": "2023-24",
                "degree": "BE",
                "departmentid": "CS",
                "termNumber": "1",
                "scheme": "Scheme 2022",
                "examid": "DECEMBER_FEBRUARY_2023-24",
                "usn": USN
            }

            x = requests.get(url=url, cookies=cookies, headers=headers, params=params)
            if(x.status_code == 200):
                save(   parse(x.json()))
                print("done "+str(i))

        except Exception:
            print("exception in i =  " + str(i))

    for i in range(10,  99):
        try:
            USN = 'GIT23CS0' + str(i) + '-T'
            url = "https://klsgroup.dhi-edu.com/dhiapiserver/api/university-exam/score/students/64e4697835a997bc3f6d305a?academicyear=2023-24&degree=BE&departmentid=CS&termNumber=1&scheme=Scheme%202022&examid=DECEMBER_FEBRUARY_2023-24&usn=" + USN
            params = {
                "academicyear": "2023-24",
                "degree": "BE",
                "departmentid": "CS",
                "termNumber": "1",
                "scheme": "Scheme 2022",
                "examid": "DECEMBER_FEBRUARY_2023-24",
                "usn": USN
            }

            x = requests.get(url=url, cookies=cookies, headers=headers, params=params)
            if(x.status_code == 200):
                save(   parse(x.json()))
                print("done "+str(i))

        except Exception:
            print("exception in i =  " + str(i))

    for i in range(100,  300):
        try:
            USN = 'GIT23CS' + str(i) + '-T'
            url = "https://klsgroup.dhi-edu.com/dhiapiserver/api/university-exam/score/students/64e4697835a997bc3f6d305a?academicyear=2023-24&degree=BE&departmentid=CS&termNumber=1&scheme=Scheme%202022&examid=DECEMBER_FEBRUARY_2023-24&usn=" + USN
            params = {
                "academicyear": "2023-24",
                "degree": "BE",
                "departmentid": "CS",
                "termNumber": "1",
                "scheme": "Scheme 2022",
                "examid": "DECEMBER_FEBRUARY_2023-24",
                "usn": USN
            }

            x = requests.get(url=url, cookies=cookies, headers=headers, params=params)
            if(x.status_code == 200):
                save(   parse(x.json()))
                print("done "+str(i))
        except Exception:
            print("exception in i =  " + str(i))
    




def create_excel_file(filename, headers):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    workbook.save(filename)

def append_row(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(filename)

def parse(res):
    result = dict()
    result['USN'] = res['usn']
    result['name'] = res['studentName']
    result['totalScore'] = 0.0
    result['totalCredit'] = 0.0
    result['scores'] = dict()
    result['grades'] = dict()
    result['credits'] = dict()

    for course in res['ueScoreGraph']:
        result['scores'][course['courseCode']]= float(course['totalScore'])
        result['credits'][course['courseCode']]= float(course['creditPoint'])
        result['totalCredit'] += float(course['creditPoint'])
        result['totalScore'] += float(course['totalScore'])

    return result

def save(data):
    global dfc, dfs
    temp = []
    temp.append(data['name'])
    temp.append(data['USN'])

    temp.append(data['totalCredit'])
    for course in data['credits']:
        temp.append(data['credits'][course])
    append_row('credits.xlsx', temp)

    temp2 = []
    temp2.append(data['name'])
    temp2.append(data['USN'])

    temp2.append(data['totalScore'])
    for course in data['scores']:
        temp2.append(data['scores'][course])
    append_row('scores.xlsx', temp2)


main()