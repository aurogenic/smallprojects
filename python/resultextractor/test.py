import json
import pandas as pd
from openpyxl import Workbook, load_workbook

# dfs = pd.DataFrame(columns=['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'])
# dfc = pd.DataFrame(columns=['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'])

file = open('result.json', 'r')
res = json.loads( file.read())
file.close

def create_excel_file(filename, headers):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    workbook.save(filename)

def append_row(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(filename)

create_excel_file("credits.xlsx", ['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'] )
create_excel_file("scores.xlsx", ['Name', 'USN', 'Total','22ETC15F', '22IIL18', '22KSK17', '22MATS11', '22PHYS12', '22POP13', '22ENG16', '22ESC143'] )

def parse(res):
    result = dict()
    result['USN'] = res['usn']
    result['name'] = res['studentName']
    result['totalScore'] = 0.0
    result['totalCredit'] = 0.0
    result['scores'] = dict()
    result['grades'] = dict()
    result['credits'] = dict()

    for course in res['ueScoreGraph']:
        result['scores'][course['courseCode']]= float(course['totalScore'])
        result['credits'][course['courseCode']]= float(course['creditPoint'])
        result['totalCredit'] += float(course['creditPoint'])
        result['totalScore'] += float(course['totalScore'])

    return result

def save(data):
    global dfc, dfs
    temp = []
    temp.append(data['name'])
    temp.append(data['USN'])

    temp.append(data['totalCredit'])
    for course in data['credits']:
        temp.append(data['credits'][course])
    append_row('credits.xlsx', temp)

    temp2 = []
    temp2.append(data['name'])
    temp2.append(data['USN'])

    temp2.append(data['totalScore'])
    for course in data['scores']:
        temp2.append(data['scores'][course])
    append_row('scores.xlsx', temp2)


data = parse(res)
save(data)
